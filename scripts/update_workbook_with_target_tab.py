import sys
import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

print("Adding dedicated High-Intent sheet to workbook...")

excel_path = r'C:\Users\User\Documents\org\ONE_BIG_WORLDWIDE_LICENSED_LEADS_2026-08-28.xlsx'
downloads_path = r'C:\Users\User\Downloads\ONE_BIG_WORLDWIDE_LICENSED_LEADS_2026-08-28.xlsx'

df_all = pd.read_excel(excel_path, sheet_name='Leads')

# Filter target high-intent categories:
# 1. Legal representation / litigation possible
# 2. Claims handling / insurance claims authorization
# 3. Payment / transaction-dispute capability
# 4. Crypto-asset services authorization
# 5. Forensic / investigative / asset-tracing service indicated

target_categories = [
    'Legal representation / litigation possible',
    'Claims handling / insurance claims authorization',
    'Payment / transaction-dispute capability',
    'Crypto-asset services authorization',
    'Forensic / investigative / asset-tracing service indicated'
]

df_target = df_all[df_all['Recovery or disputes relevance'].isin(target_categories)].copy()

# Sort target leads: High priority first, then country, then company
priority_order = {'High': 0, 'Medium': 1, 'Review': 2}
df_target['prio_rank'] = df_target['Priority'].map(priority_order).fillna(3)
df_target = df_target.sort_values(by=['prio_rank', 'Country', 'Company name']).drop(columns=['prio_rank'])

print(f"Target high-intent leads count: {len(df_target)}")

# Rebuild workbook with 3 sheets:
# 1. Read me
# 2. Direct Recovery & Litigation (Targeted 888 leads)
# 3. All Leads (15,974 leads)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# 1. Read me Sheet
ws_readme = wb.create_sheet(title="Read me")
readme_content = [
    ["WORLDWIDE LICENSED ORGANIZATIONS — WEBSITE DEVELOPMENT LEADS"],
    ["Dataset Version: 2026-08-28 | Consolidated Single-Workbook Edition"],
    [""],
    ["1. COMMERCIAL OBJECTIVE & STRUCTURE"],
    ["This workbook contains regulated organizations worldwide that have NO working website recorded"],
    ["in official directories, making them prime prospects for website development services."],
    [""],
    ["Workbook Structure:"],
    ["  * Tab 'Direct Recovery & Litigation': 888 top-tier organizations with explicit authorization for:"],
    ["      - Legal representation / court litigation (Law firms, Solicitors, Advocates);"],
    ["      - Claims handling / Chargebacks / Financial disputes (Claims Management Companies, AFSL);"],
    ["      - Crypto-asset services / Blockchain forensics (VASPs, CASPs, Tracing specialists)."],
    ["  * Tab 'All Leads': Complete worldwide database of 15,974 licensed organizations across 29 countries."],
    [""],
    ["2. REGULATORY REALITY REGARDING 'GUARANTEED RECOVERY'"],
    ["In regulated jurisdictions, claiming 'guaranteed fund recovery' is illegal or considered a fraud indicator."],
    ["However, the organizations in this database hold legitimate official licenses to perform:"],
    ["  - Legal court representation, freezing orders (injunctions), and fraud litigation;"],
    ["  - Regulated claims management, banking dispute resolution, and insurance settlements;"],
    ["  - Blockchain transaction tracing and forensic crypto analysis."],
    [""],
    ["3. HOW TO PITCH WEBSITE DEVELOPMENT SERVICES"],
    ["- Emphasize trust, regulatory compliance, and professional client intake portals."],
    ["- Highlight that lack of a modern, compliant website causes prospective clients to doubt credibility."]
]

for r_idx, row in enumerate(readme_content, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws_readme.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(name='Calibri', size=14, bold=True, color='1F497D')
        elif r_idx == 2:
            cell.font = Font(name='Calibri', size=11, italic=True, color='595959')
        elif str(val).startswith(('1.', '2.', '3.')):
            cell.font = Font(name='Calibri', size=11, bold=True, color='002060')
        else:
            cell.font = Font(name='Calibri', size=10)

ws_readme.column_dimensions['A'].width = 115

# Styling Definitions
header_fill_target = PatternFill(start_color="004C6D", end_color="004C6D", fill_type="solid")
header_fill_all = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)

min_max_widths = {
    'Priority': 12,
    'Country': 16,
    'Company name': 35,
    'Organization type': 30,
    'Regulator or professional body': 35,
    'Licence / register number': 25,
    'Licence or registration status': 32,
    'Status checked date': 18,
    'Source date': 18,
    'Licence scope / permitted services': 45,
    'Recovery or disputes relevance': 35,
    'Website URL': 22,
    'Website status': 32,
    'Website verification method': 35,
    'Email': 25,
    'Phone': 20,
    'City / address': 28,
    'Official source URL': 40,
    'Source file or dataset': 38,
    'Confidence': 14,
    'Next verification action': 45
}

def style_sheet(ws, df, header_fill):
    headers = list(df.columns)
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    data_rows = df.values.tolist()
    for r_idx, r_data in enumerate(data_rows, start=2):
        ws.append(r_data)
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=r_idx, column=col_idx)
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border
            if col_idx in [1, 2, 7, 8, 9, 13, 20]:
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.alignment = Alignment(horizontal="left", vertical="top")
                
    for col_num, col_name in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        w = min_max_widths.get(col_name, 25)
        ws.column_dimensions[col_letter].width = w

    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(df) + 1}"
    ws.freeze_panes = "A2"

# 2. Create Direct Recovery & Litigation Sheet
ws_target = wb.create_sheet(title="Direct Recovery & Litigation")
style_sheet(ws_target, df_target, header_fill_target)

# 3. Create All Leads Sheet
ws_all = wb.create_sheet(title="All Leads")
style_sheet(ws_all, df_all, header_fill_all)

wb.save(excel_path)
wb.save(downloads_path)

print(f"Updated workbook saved successfully to:")
print(f"  1. {excel_path}")
print(f"  2. {downloads_path}")
